import re
import io
import pandas as pd
from openpyxl import load_workbook

# TODO упростить
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 МБ, можно поменять под свои нужды
MAX_ROWS = 500

def xlsx_to_html(xlsx_bytes: bytes, include_formulas=True, include_comments=True) -> str:
    """
    Форматирует xlsx в html
    """
    if len(xlsx_bytes) > MAX_FILE_SIZE:
        raise ValueError(f"Файл слишком большой (> {MAX_FILE_SIZE} байт)")

    parts: list[str] = []

    xls = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")

    if not xls.sheet_names:
        return "<p>Файл не содержит листов</p>"

    first_sheet = xls.sheet_names[0]

    df = xls.parse(sheet_name=first_sheet, dtype=str, nrows=MAX_ROWS)
    df = df.fillna("")

    parts.append(f"<h2>{first_sheet}</h2>")
    parts.append(df.to_html(index=False, border=1, escape=True))

    if include_formulas or include_comments:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False, read_only=True)
        ws = wb.worksheets[0]  # только первый лист

        if include_formulas:
            formula_items = []
            try:
                for row in ws.iter_rows(max_row=MAX_ROWS):
                    for c in row:
                        val = c.value
                        # формула — строка, начинающаяся с '='
                        if isinstance(val, str) and val.startswith("="):
                            formula_items.append(
                                f"<tr><td>{ws.title}</td>"
                                f"<td>{c.coordinate}</td>"
                                f"<td><code>{val}</code></td></tr>"
                            )
            except Exception:
                pass

            if formula_items:
                parts.append("<h3>Формулы</h3>")
                parts.append(
                    "<table border='1'>"
                    "<tr><th>Лист</th><th>Адрес</th><th>Формула</th></tr>"
                )
                parts.extend(formula_items)
                parts.append("</table>")

        if include_comments:
            comment_items = []
            try:
                for row in ws.iter_rows(max_row=MAX_ROWS):
                    for c in row:
                        comment = c.comment
                        if comment:
                            txt = (comment.text or "").replace("\n", "<br>")
                            comment_items.append(
                                f"<tr><td>{ws.title}</td>"
                                f"<td>{c.coordinate}</td>"
                                f"<td>{txt}</td></tr>"
                            )
            except Exception:
                pass

            if comment_items:
                parts.append("<h3>Комментарии</h3>")
                parts.append(
                    "<table border='1'>"
                    "<tr><th>Лист</th><th>Адрес</th><th>Комментарий</th></tr>"
                )
                parts.extend(comment_items)
                parts.append("</table>")

    html = "\n".join(parts)
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", html)
